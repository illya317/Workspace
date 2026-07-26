#!/usr/bin/env python3
"""Convert one immutable Library version into Markdown and locator-rich layout JSON."""

from __future__ import annotations

import argparse
from datetime import date, datetime
import hashlib
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def artifact(path: Path, kind: str, mime_type: str) -> dict[str, Any]:
    return {
        "kind": kind,
        "fileName": path.name,
        "mimeType": mime_type,
        "fileSizeBytes": path.stat().st_size,
        "checksumSha256": sha256(path),
    }


def pdf_has_usable_text(input_path: Path) -> tuple[bool, int]:
    from pypdf import PdfReader

    reader = PdfReader(str(input_path))
    chars = sum(len((page.extract_text() or "").strip()) for page in reader.pages)
    threshold = max(80, len(reader.pages) * 30)
    return chars >= threshold, len(reader.pages)


def native_pdf(input_path: Path) -> tuple[str, dict[str, Any], int]:
    import pdfplumber

    markdown: list[str] = []
    segments: list[dict[str, Any]] = []
    with pdfplumber.open(input_path) as pdf:
        for index, page in enumerate(pdf.pages, start=1):
            text = (page.extract_text(x_tolerance=2, y_tolerance=3) or "").strip()
            markdown.extend([f"## Page {index}", "", text or "<!-- no extractable text -->", ""])
            if text:
                segments.append({
                    "ordinal": len(segments),
                    "kind": "page_text",
                    "text": text,
                    "locator": {"schemaVersion": "v1", "page": index},
                })
        return "\n".join(markdown).strip() + "\n", {
            "schemaVersion": "v1",
            "engine": "pdfplumber",
            "segments": segments,
        }, len(pdf.pages)


def tesseract_pdf(input_path: Path) -> tuple[str, dict[str, Any], int]:
    markdown: list[str] = []
    segments: list[dict[str, Any]] = []
    with tempfile.TemporaryDirectory(prefix="library-ocr-") as directory:
        prefix = Path(directory) / "page"
        subprocess.run(
            ["pdftoppm", "-jpeg", "-r", "150", str(input_path), str(prefix)],
            check=True,
            capture_output=True,
            text=True,
        )
        pages = sorted(Path(directory).glob("page-*.jpg"))
        if not pages:
            raise RuntimeError("pdftoppm produced no OCR pages")
        for index, page in enumerate(pages, start=1):
            completed = subprocess.run(
                [
                    "env", "OMP_THREAD_LIMIT=1", "tesseract", str(page), "stdout",
                    "-l", "chi_sim+eng", "--psm", "3",
                ],
                check=True,
                capture_output=True,
                text=True,
            )
            text = completed.stdout.strip()
            markdown.extend([f"## Page {index}", "", text or "<!-- no extractable text -->", ""])
            if text:
                segments.append({
                    "ordinal": len(segments),
                    "kind": "page_ocr",
                    "text": text,
                    "locator": {"schemaVersion": "v1", "page": index},
                })
    return "\n".join(markdown).strip() + "\n", {
        "schemaVersion": "v1",
        "engine": "tesseract",
        "ocrUsed": True,
        "segments": segments,
    }, len(pages)


def excel_column_name(index: int) -> str:
    """Convert a one-based column index to an Excel column name."""
    value = index
    parts: list[str] = []
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        parts.append(chr(ord("A") + remainder))
    return "".join(reversed(parts))


def docling_table_segments(exported: dict[str, Any], start_ordinal: int) -> list[dict[str, Any]]:
    """Expose spreadsheet table rows with stable sheet/cell locators."""
    sheet_names = {
        str(group.get("self_ref")): str(group.get("name"))
        for group in exported.get("groups", [])
        if group.get("label") == "sheet" and group.get("self_ref") and group.get("name")
    }
    segments: list[dict[str, Any]] = []
    for table_index, table in enumerate(exported.get("tables", []), start=1):
        parent = table.get("parent") or {}
        provenance = table.get("prov") or []
        first_provenance = provenance[0] if provenance else {}
        sheet = sheet_names.get(str(parent.get("$ref")))
        if not sheet:
            sheet = f"Sheet {first_provenance.get('page_no') or table_index}"

        rows: dict[int, list[dict[str, Any]]] = {}
        headers: dict[int, str] = {}
        for cell in (table.get("data") or {}).get("table_cells", []):
            text = " ".join(str(cell.get("text") or "").split())
            if not text:
                continue
            row_index = int(cell.get("start_row_offset_idx") or 0)
            column_index = int(cell.get("start_col_offset_idx") or 0)
            rows.setdefault(row_index, []).append({**cell, "normalized_text": text})
            if cell.get("column_header"):
                headers[column_index] = text

        for row_index in sorted(rows):
            cells = sorted(rows[row_index], key=lambda item: int(item.get("start_col_offset_idx") or 0))
            row_is_header = all(bool(cell.get("column_header")) for cell in cells)
            values: list[str] = []
            for cell in cells:
                text = str(cell["normalized_text"])
                column_index = int(cell.get("start_col_offset_idx") or 0)
                header = headers.get(column_index)
                values.append(text if row_is_header or not header else f"{header}: {text}")
            start_column = min(int(cell.get("start_col_offset_idx") or 0) for cell in cells) + 1
            end_column = max(int(cell.get("end_col_offset_idx") or 1) for cell in cells)
            cell_range = (
                f"{excel_column_name(start_column)}{row_index + 1}:"
                f"{excel_column_name(end_column)}{row_index + 1}"
            )
            segments.append({
                "ordinal": start_ordinal + len(segments),
                "kind": "table_row",
                "text": " | ".join(values),
                "locator": {
                    "schemaVersion": "v1",
                    "sheet": sheet,
                    "cellRange": cell_range,
                },
            })
    return segments


def spreadsheet_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return " ".join(str(value).split())


def spreadsheet_segments(input_path: Path) -> list[dict[str, Any]]:
    """Read XLSX rows directly so locators remain tied to the original workbook."""
    from openpyxl import load_workbook

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    segments: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            headers: dict[int, str] = {}
            header_row_number: int | None = None
            for row in worksheet.iter_rows():
                populated = [(cell.column, spreadsheet_value(cell.value)) for cell in row if spreadsheet_value(cell.value)]
                if not populated:
                    continue
                if header_row_number is None:
                    header_row_number = row[0].row
                    headers = dict(populated)
                is_header = row[0].row == header_row_number
                values = [
                    text if is_header or not headers.get(column) else f"{headers[column]}: {text}"
                    for column, text in populated
                ]
                start_column = min(column for column, _text in populated)
                end_column = max(column for column, _text in populated)
                segments.append({
                    "ordinal": len(segments),
                    "kind": "table_row",
                    "text": " | ".join(values),
                    "locator": {
                        "schemaVersion": "v1",
                        "sheet": worksheet.title,
                        "cellRange": (
                            f"{excel_column_name(start_column)}{row[0].row}:"
                            f"{excel_column_name(end_column)}{row[0].row}"
                        ),
                    },
                })
    finally:
        workbook.close()
    return segments


def docling_convert(input_path: Path, ocr_used: bool) -> tuple[str, dict[str, Any], int]:
    from docling.document_converter import DocumentConverter

    conversion = DocumentConverter().convert(str(input_path))
    document = conversion.document
    exported = document.export_to_dict()
    if input_path.suffix.lower() in {".xlsx", ".xlsm"}:
        segments = spreadsheet_segments(input_path)
    else:
        segments = []
        for item in exported.get("texts", []):
            text = str(item.get("text") or "").strip()
            if not text:
                continue
            provenance = item.get("prov") or []
            first = provenance[0] if provenance else {}
            locator: dict[str, Any] = {"schemaVersion": "v1"}
            if first.get("page_no"):
                locator["page"] = int(first["page_no"])
            if len(locator) == 1:
                locator["sectionPath"] = [str(item.get("label") or "document")]
            segments.append({
                "ordinal": len(segments),
                "kind": str(item.get("label") or "text"),
                "text": text,
                "locator": locator,
            })
        segments.extend(docling_table_segments(exported, len(segments)))
    layout = {
        "schemaVersion": "v1",
        "engine": "docling",
        "ocrUsed": ocr_used,
        "conversionStatus": str(conversion.status),
        "segments": segments,
        "doclingDocument": exported,
    }
    return document.export_to_markdown(), layout, len(conversion.pages)


def plain_text_convert(input_path: Path) -> tuple[str, dict[str, Any], int]:
    text = input_path.read_text(encoding="utf-8", errors="replace")
    return text, {
        "schemaVersion": "v1",
        "engine": "plain-text",
        "segments": [{
            "ordinal": 0,
            "kind": "text",
            "text": text,
            "locator": {"schemaVersion": "v1", "sectionPath": [input_path.name]},
        }],
    }, 1


def process(input_path: Path) -> tuple[str, dict[str, Any], int, bool, str]:
    suffix = input_path.suffix.lower()
    if suffix == ".pdf":
        has_text, _ = pdf_has_usable_text(input_path)
        if has_text:
            markdown, layout, pages = native_pdf(input_path)
            return markdown, layout, pages, False, "pdfplumber"
        markdown, layout, pages = tesseract_pdf(input_path)
        return markdown, layout, pages, True, "tesseract"
    if suffix in {".md", ".txt", ".csv", ".json", ".xml"}:
        markdown, layout, pages = plain_text_convert(input_path)
        return markdown, layout, pages, False, "plain-text"
    markdown, layout, pages = docling_convert(input_path, suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff"})
    return markdown, layout, pages, layout["ocrUsed"], "docling"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--document-uid", required=True)
    parser.add_argument("--version-uid", required=True)
    parser.add_argument("--input-checksum", required=True)
    parser.add_argument("--pipeline-version", default="v1.0.4")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    output_dir = Path(args.output_dir).resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"input file missing: {input_path}")
    if sha256(input_path) != args.input_checksum:
        raise ValueError("input checksum mismatch")
    if output_dir.exists():
        raise FileExistsError(f"output directory already exists: {output_dir}")

    output_dir.parent.mkdir(parents=True, exist_ok=True)
    temporary_dir = Path(tempfile.mkdtemp(prefix=f".{output_dir.name}-", dir=output_dir.parent))
    try:
        markdown, layout, page_count, ocr_used, engine = process(input_path)
        layout.update({
            "documentUid": args.document_uid,
            "versionUid": args.version_uid,
            "inputChecksumSha256": args.input_checksum,
            "pipelineVersion": args.pipeline_version,
            "pageCount": page_count,
        })
        markdown_path = temporary_dir / "document.md"
        layout_path = temporary_dir / "layout.json"
        markdown_path.write_text(markdown, encoding="utf-8")
        layout_path.write_text(json.dumps(layout, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        warnings = []
        if ocr_used:
            warnings.append("ocr_requires_review")
        if not markdown.strip():
            warnings.append("empty_markdown")
        result = {
            "status": "warning" if warnings else "succeeded",
            "engine": engine,
            "ocrUsed": ocr_used,
            "pageCount": page_count,
            "segmentCount": len(layout["segments"]),
            "artifacts": [
                artifact(markdown_path, "markdown", "text/markdown"),
                artifact(layout_path, "layout-json", "application/json"),
            ],
            "warnings": warnings,
        }
        (temporary_dir / "result.json").write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        temporary_dir.rename(output_dir)
        print(json.dumps(result, ensure_ascii=False))
    except Exception:
        shutil.rmtree(temporary_dir, ignore_errors=True)
        raise


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # noqa: BLE001 - CLI must emit one stable failure line
        print(json.dumps({"status": "failed", "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        raise
